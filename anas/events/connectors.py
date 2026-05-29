from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import json
import re

from openpyxl import load_workbook
from pypdf import PdfReader
import requests

from anas.events.html_text import html_to_text
from anas.events.schemas import SeedEvent


@dataclass(frozen=True)
class FetchResult:
    http_status: int
    final_url: str
    raw_html: str
    raw_text: str


class BaseOfficialConnector:
    def __init__(self, *, timeout: int = 45):
        self.timeout = timeout

    def fetch(self, event: SeedEvent) -> FetchResult:
        response = requests.get(
            event.source_url,
            headers=browser_headers(),
            timeout=self.timeout,
            allow_redirects=True,
        )
        response.raise_for_status()
        raw_text = response_to_text(response, event=event)
        return FetchResult(
            http_status=response.status_code,
            final_url=response.url,
            raw_html=response_to_raw_payload(response),
            raw_text=raw_text,
        )


class BLSConnector(BaseOfficialConnector):
    """BLS archives can reject simple scripted requests; try official fallbacks."""

    def fetch(self, event: SeedEvent) -> FetchResult:
        urls = [event.source_url]
        if event.source_url.endswith(".htm"):
            urls.append(event.source_url[:-4] + ".txt")
        if event.source_url.endswith(".html"):
            urls.append(event.source_url[:-5] + ".txt")

        last_status: int | str = "network"
        for url in urls:
            try:
                response = requests.get(
                    url,
                    headers=bls_headers(),
                    timeout=self.timeout,
                    allow_redirects=True,
                )
                last_status = response.status_code
                response.raise_for_status()
            except requests.RequestException:
                continue

            content_type = response.headers.get("content-type", "").lower()
            raw = response.text
            raw_text = raw if "text/plain" in content_type or url.endswith(".txt") else response_to_text(response, event=event)
            return FetchResult(
                http_status=response.status_code,
                final_url=response.url,
                raw_html=response_to_raw_payload(response),
                raw_text=raw_text,
            )

        if event.document_type == "cpi_release":
            return self._fetch_cpi_api_fallback(event, last_status=last_status)

        raise RuntimeError(f"BLS request failed for event_id={event.event_id}, status={last_status}")

    def _fetch_cpi_api_fallback(self, event: SeedEvent, *, last_status: int | str) -> FetchResult:
        target_year, target_month = _extract_month_year(event.title)
        if target_year is None or target_month is None:
            raise RuntimeError(f"BLS request failed for event_id={event.event_id}, status={last_status}")

        series = {
            "CUSR0000SA0": "CPI-U all items, seasonally adjusted",
            "CUUR0000SA0": "CPI-U all items, not seasonally adjusted",
            "CUSR0000SA0L1E": "CPI-U all items less food and energy, seasonally adjusted",
            "CUUR0000SA0L1E": "CPI-U all items less food and energy, not seasonally adjusted",
        }
        api_payload = {}
        for series_id, label in series.items():
            url = (
                "https://api.bls.gov/publicAPI/v2/timeseries/data/"
                f"{series_id}?startyear={target_year - 1}&endyear={target_year}"
            )
            response = requests.get(url, headers=browser_headers(), timeout=self.timeout)
            response.raise_for_status()
            api_payload[series_id] = response.json()

        text = _build_cpi_api_text(
            event=event,
            target_year=target_year,
            target_month=target_month,
            series_labels=series,
            api_payload=api_payload,
        )
        return FetchResult(
            http_status=200,
            final_url="https://api.bls.gov/publicAPI/v2/timeseries/data/",
            raw_html=json.dumps(api_payload, ensure_ascii=False),
            raw_text=text,
        )


def connector_for_event(event: SeedEvent, *, timeout: int = 45) -> BaseOfficialConnector:
    if event.source_policy_id == "bls_public_domain" or "bls.gov" in event.source_url:
        return BLSConnector(timeout=timeout)
    return BaseOfficialConnector(timeout=timeout)


def response_to_text(response: requests.Response, *, event: SeedEvent) -> str:
    content_type = response.headers.get("content-type", "").lower()
    url = response.url.lower()
    content = response.content

    if "application/pdf" in content_type or url.endswith(".pdf"):
        return pdf_bytes_to_text(content, title=event.title, source_url=event.source_url)

    if (
        "spreadsheet" in content_type
        or "excel" in content_type
        or url.endswith(".xlsx")
        or url.endswith(".xlsm")
    ):
        return xlsx_bytes_to_text(content, title=event.title, source_url=event.source_url)

    if "text/plain" in content_type or url.endswith(".txt"):
        return response.text

    return html_to_text(response.text)


def response_to_raw_payload(response: requests.Response) -> str:
    content_type = response.headers.get("content-type", "").lower()
    url = response.url.lower()
    if (
        "application/pdf" in content_type
        or "spreadsheet" in content_type
        or "excel" in content_type
        or url.endswith((".pdf", ".xlsx", ".xlsm"))
    ):
        return response.content.hex()
    return response.text


def pdf_bytes_to_text(content: bytes, *, title: str, source_url: str) -> str:
    reader = PdfReader(BytesIO(content))
    lines = [
        title,
        f"Official PDF source: {source_url}",
        "",
    ]
    for page_number, page in enumerate(reader.pages, 1):
        text = page.extract_text() or ""
        if text.strip():
            lines.append(f"[PDF page {page_number}]")
            lines.append(text.strip())
            lines.append("")
    return "\n".join(lines).strip()


def xlsx_bytes_to_text(content: bytes, *, title: str, source_url: str, max_rows_per_sheet: int = 250) -> str:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    lines = [
        title,
        f"Official spreadsheet source: {source_url}",
        "",
    ]
    for sheet in workbook.worksheets:
        lines.append(f"[XLSX sheet: {sheet.title}]")
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_index > max_rows_per_sheet:
                lines.append(f"... truncated after {max_rows_per_sheet} rows")
                break
            values = [_format_cell(value) for value in row]
            values = [value for value in values if value]
            if values:
                lines.append(" | ".join(values))
        lines.append("")
    return "\n".join(lines).strip()


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.10g}"
    return str(value).strip()


def browser_headers() -> dict[str, str]:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/125.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,text/plain;q=0.8,*/*;q=0.7",
        "Accept-Language": "en-US,en;q=0.9",
        "Connection": "keep-alive",
    }


def bls_headers() -> dict[str, str]:
    headers = browser_headers()
    headers.update(
        {
            "Host": "www.bls.gov",
            "Referer": "https://www.bls.gov/news.release/cpi.htm",
        }
    )
    return headers


def _extract_month_year(title: str) -> tuple[int | None, int | None]:
    month_lookup = {
        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12,
    }
    match = re.search(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})",
        title,
        re.IGNORECASE,
    )
    if not match:
        return None, None
    return int(match.group(2)), month_lookup[match.group(1).lower()]


def _build_cpi_api_text(
    *,
    event: SeedEvent,
    target_year: int,
    target_month: int,
    series_labels: dict[str, str],
    api_payload: dict,
) -> str:
    period = f"M{target_month:02d}"
    prior_month = target_month - 1
    prior_year = target_year
    if prior_month == 0:
        prior_month = 12
        prior_year -= 1
    prior_period = f"M{prior_month:02d}"

    lines = [
        f"{event.title}",
        "Official BLS Public Data API fallback record.",
        f"Original archived news release URL: {event.source_url}",
        "The HTML/PDF archive page rejected local scripted access, so ANAS used the official BLS Public Data API.",
        f"Target CPI reference period: {target_year}-{target_month:02d}.",
        "",
    ]

    for series_id, label in series_labels.items():
        values = _series_values(api_payload[series_id])
        current = values.get((target_year, period))
        previous = values.get((prior_year, prior_period))
        year_ago = values.get((target_year - 1, period))
        lines.append(f"Series {series_id}: {label}.")
        if current is not None:
            lines.append(f"Index value for {period} {target_year}: {current:.3f}.")
        if current is not None and previous is not None:
            mom = (current - previous) / previous * 100
            lines.append(f"One-month percent change from previous month: {mom:.3f} percent.")
        if current is not None and year_ago is not None:
            yoy = (current - year_ago) / year_ago * 100
            lines.append(f"Twelve-month percent change from year earlier: {yoy:.3f} percent.")
        lines.append("")

    lines.extend(
        [
            "Interpretation boundary: this fallback text is built only from official BLS API data and seed metadata.",
            "It is suitable for the ANAS demo as an official-source inflation event record.",
            "The exposure extractor should treat it as an inflation and monetary-policy event, not as a copyrighted news article.",
        ]
    )
    return "\n".join(lines)


def _series_values(payload: dict) -> dict[tuple[int, str], float]:
    values = {}
    for series in payload.get("Results", {}).get("series", []):
        for item in series.get("data", []):
            values[(int(item["year"]), item["period"])] = float(item["value"])
    return values
